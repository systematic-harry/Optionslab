# ============================================================
#  load_universe.py
#  OptionLab — Universe Loader
#
#  What it does:
#  1. Downloads nifty50_universe.xlsx from GCS
#  2. For each stock, looks up conId from IB TWS
#  3. Fills the ConId column in Excel
#  4. Uploads updated Excel back to GCS
#
#  Run once in Spyder:
#  from load_universe import run
#  run()
# ============================================================

import io
import time
import pandas as pd
import openpyxl
from ibapi.contract import Contract
from ib_core import (
    get_app, is_connected, connect_local,
    req_manager, wait_for,
    get_bucket
)

# ── Config ────────────────────────────────────────────────────
UNIVERSE_FILE   = "nifty50_universe.xlsx"
UNIVERSE_FOLDER = "universe"


# ── Download Excel from GCS ───────────────────────────────────
def download_universe():
    print("\n  Downloading universe file from GCS...")
    bucket = get_bucket()
    path   = f"{UNIVERSE_FOLDER}/{UNIVERSE_FILE}"
    blob   = bucket.blob(path)

    if not blob.exists():
        print(f"  [Error] File not found: {path}")
        return None

    data = blob.download_as_bytes()
    df   = pd.read_excel(io.BytesIO(data))
    print(f"  Downloaded — {len(df)} stocks found")
    return df


# ── Upload Excel back to GCS ──────────────────────────────────
def upload_universe(df):
    print("\n  Uploading updated universe to GCS...")
    bucket = get_bucket()
    path   = f"{UNIVERSE_FOLDER}/{UNIVERSE_FILE}"
    blob   = bucket.blob(path)

    # Save DataFrame back to Excel in memory
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Nifty50")

        # Style the sheet
        ws = writer.sheets["Nifty50"]
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color="1F3864",
                                   end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    buffer.seek(0)
    blob.upload_from_string(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    print(f"  Uploaded -> gs://option_lab_data/{path}")


# ── Lookup conId from IB ──────────────────────────────────────
def lookup_conid(app, symbol, exchange="NSE", currency="INR"):
    """
    Request contract details from IB for a symbol.
    Returns conId (int) or None if not found.
    """
    rid      = req_manager.next(f"{symbol}_CONTRACT")
    contract = Contract()
    contract.symbol   = symbol
    contract.secType  = "STK"
    contract.exchange = exchange
    contract.currency = currency

    app.reqContractDetails(rid, contract)
    success = wait_for(app, rid, timeout=10)

    if not success:
        return None

    details = app.contract_map.get(rid)
    if details:
        return details.get("conid") or details.get("con_id")

    return None


# ── Main ──────────────────────────────────────────────────────
def run():
    print("\n" + "="*50)
    print("  OPTIONLAB — Load Universe")
    print("="*50)

    # Connect if not already
    if not is_connected():
        connect_local()

    app = get_app()
    if app is None:
        print("  [Error] Not connected to TWS.")
        return

    # Download universe Excel from GCS
    df = download_universe()
    if df is None:
        return

    # Ensure ConId column exists
    if "ConId" not in df.columns:
        df["ConId"] = None

    print(f"\n  Looking up conIds for {len(df)} stocks...\n")

    success_count = 0
    failed        = []

    for i, row in df.iterrows():
        symbol   = str(row["IB_Symbol"]).strip()
        name     = str(row["Name"]).strip()
        exchange = str(row.get("Exchange", "NSE")).strip()
        currency = str(row.get("Currency", "INR")).strip()

        # Skip if already has conId
        existing = row.get("ConId")
        if pd.notna(existing) and str(existing).strip() not in ["", "None", "nan"]:
            print(f"  [{i+1:02d}] {name} — already has conId: {existing}")
            success_count += 1
            continue

        print(f"  [{i+1:02d}] Looking up {name} ({symbol})...", end=" ")

        conid = lookup_conid(app, symbol, exchange, currency)

        if conid:
            df.at[i, "ConId"] = conid
            print(f"conId: {conid}")
            success_count += 1
        else:
            print("NOT FOUND")
            failed.append((name, symbol))

        time.sleep(0.5)  # avoid pacing violations

    # Upload updated Excel back to GCS
    upload_universe(df)

    # Summary
    print(f"\n{'='*50}")
    print(f"  Done!")
    print(f"  Success : {success_count}/{len(df)}")
    if failed:
        print(f"  Failed  : {len(failed)} stocks")
        for name, symbol in failed:
            print(f"    - {name} ({symbol})")
        print("\n  Tip: Check IB_Symbol column for these stocks")
        print("  Use IB Contract Search: https://www.interactivebrokers.com/en/trading/symbol-search.php")
    print(f"{'='*50}\n")

    return df


# ── Also expose as helper to get universe from cloud ─────────
def get_universe():
    """
    Download universe Excel from GCS and return as DataFrame.
    Use this in screener.py, fetch_data.py etc.
    """
    df = download_universe()
    if df is None:
        return {}

    # Return as dict: {symbol: conId}
    universe = {}
    for _, row in df.iterrows():
        symbol = str(row["IB_Symbol"]).strip()
        conid  = row.get("ConId")
        if pd.notna(conid) and str(conid).strip() not in ["", "None", "nan"]:
            universe[symbol] = int(conid)

    print(f"  Universe loaded — {len(universe)} stocks with conIds")
    return universe


# ── Run ───────────────────────────────────────────────────────
if __name__ == "__main__":
    run()
